import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.styles import ParagraphStyle

from tools.getter import get_daily_weather_data
from tools.period_aggregator import extract_daily_data

# шрифты:
import os
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping

# Регистрируем шрифты
FONT_REGULAR = os.path.join("fonts", "DejaVuSans.ttf")
FONT_BOLD = os.path.join("fonts", "DejaVuSans-Bold.ttf")
pdfmetrics.registerFont(TTFont('DejaVu', FONT_REGULAR))
pdfmetrics.registerFont(TTFont('DejaVu-Bold', FONT_BOLD))
addMapping('DejaVu', 0, 0, 'DejaVu')  # normal
addMapping('DejaVu', 1, 0, 'DejaVu-Bold')  # bold
RUSSIAN_FONT = 'DejaVu'


def get_period_data(year: int, month: int = None, quarter: int = None) -> List[Dict[str, Any]]:
    """
    Получает данные за указанный период в дневном формате
    """
    if month:
        # Месяц
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
    elif quarter:
        # Квартал
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)

        if quarter == 4:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_month = start_month + 2
            end_date = datetime(year, end_month + 1, 1) - timedelta(days=1)
    else:
        # Год
        start_date = datetime(year, 1, 1)
        end_date = datetime(year, 12, 31)

    # Получаем данные
    weather_data = get_daily_weather_data(start_date, end_date)
    if not weather_data:
        raise Exception("Не удалось получить данные")

    # Извлекаем данные
    daily_data = extract_daily_data(weather_data)

    return daily_data


def generate_excel_bytes(data: List[Dict[str, Any]], period_name: str) -> io.BytesIO:
    """
    Генерирует Excel файл в памяти и возвращает BytesIO
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Погода {period_name}"

    # Заголовки
    headers = [
        "Дата", "Температура (°C)", "Влажность (%)",
        "Осадки (мм)", "Ветер (м/с)", "Порывы (м/с)",
        "Облачность (%)", "Солнце (ч)"
    ]

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Заголовок
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Данные
    for row, day in enumerate(data, 2):
        ws.cell(row=row, column=1, value=day["date"]).border = thin_border
        ws.cell(row=row, column=2, value=round(day["avg_temp"], 1)).border = thin_border
        ws.cell(row=row, column=3, value=round(day["avg_humidity"])).border = thin_border
        ws.cell(row=row, column=4, value=round(day["total_precipitation"], 1)).border = thin_border
        ws.cell(row=row, column=5, value=round(day["max_wind"], 1)).border = thin_border
        ws.cell(row=row, column=6, value=round(day["max_gust"], 1)).border = thin_border
        ws.cell(row=row, column=7, value=round(day["avg_cloud"])).border = thin_border
        ws.cell(row=row, column=8, value=round(day["total_sun_hours"], 1)).border = thin_border

    # Автоширина
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Сохраняем в BytesIO
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes


def generate_pdf_bytes(data: List[Dict[str, Any]], period_name: str) -> io.BytesIO:
    """
    Генерирует PDF файл в памяти и возвращает BytesIO
    """
    pdf_bytes = io.BytesIO()

    doc = SimpleDocTemplate(
        pdf_bytes,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()

    # Переопределяем стандартные стили на русский шрифт
    styles['Title'].fontName = RUSSIAN_FONT
    styles['Normal'].fontName = RUSSIAN_FONT

    # Создаем стиль для жирного текста
    styles.add(ParagraphStyle(
        name='RussianBold',
        fontName=RUSSIAN_FONT,
        fontSize=10,
        leading=12,
        alignment=0,
        spaceAfter=6
    ))

    title_style = styles['Title']
    title_style.alignment = 1  # Center

    elements = []

    # Заголовок
    elements.append(Paragraph(f"Погода в Тамани - {period_name}", title_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Таблица
    table_data = []

    # Заголовки таблицы
    headers = [
        "Дата", "Температура\n°C", "Влажность\n%", "Осадки\nмм",
        "Ветер\nм/с", "Порывы\nм/с", "Облачность\n%", "Солнце\nч"
    ]
    table_data.append(headers)

    # Данные
    for day in data:
        row = [
            day["date"][:5],  # только день.месяц
            f"{day['avg_temp']:.1f}",
            f"{day['avg_humidity']:.0f}",
            f"{day['total_precipitation']:.1f}",
            f"{day['max_wind']:.1f}",
            f"{day['max_gust']:.1f}",
            f"{day['avg_cloud']:.0f}",
            f"{day['total_sun_hours']:.1f}"
        ]
        table_data.append(row)

    # Стиль таблицы с русским шрифтом
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), RUSSIAN_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTNAME', (0, 1), (-1, -1), RUSSIAN_FONT),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)

    # Итоги
    elements.append(Spacer(1, 0.2 * inch))

    total_days = len(data)
    avg_temp = sum(d['avg_temp'] for d in data) / total_days
    total_precip = sum(d['total_precipitation'] for d in data)
    avg_cloud = sum(d['avg_cloud'] for d in data) / total_days

    summary = Paragraph(
        f"<b>Всего дней: {total_days} | "
        f"Ср. температура: {avg_temp:.1f}°C | "
        f"Осадков за период: {total_precip:.1f} мм | "
        f"Ср. облачность: {avg_cloud:.0f}%</b>",
        styles['RussianBold']  # используем жирный стиль
    )
    elements.append(summary)

    doc.build(elements)
    pdf_bytes.seek(0)
    return pdf_bytes


def generate_report_files(year: int, month: int = None, quarter: int = None) -> Tuple[io.BytesIO, io.BytesIO, str]:
    """
    Генерирует оба файла (Excel и PDF) в памяти и возвращает BytesIO объекты
    """
    if month:
        month_names = [
            "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
            "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
        ]
        period_name = f"{month_names[month - 1]} {year}"
        data = get_period_data(year, month=month)
    elif quarter:
        quarter_names = ["I квартал", "II квартал", "III квартал", "IV квартал"]
        period_name = f"{quarter_names[quarter - 1]} {year}"
        data = get_period_data(year, quarter=quarter)
    else:
        period_name = f"{year} год"
        data = get_period_data(year)

    excel_bytes = generate_excel_bytes(data, period_name)
    pdf_bytes = generate_pdf_bytes(data, period_name)

    return excel_bytes, pdf_bytes, period_name